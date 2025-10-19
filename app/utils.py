import os
from werkzeug.utils import secure_filename
from datetime import datetime
from flask import current_app
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

# ============ GESTIÓN DE ARCHIVOS ============

def allowed_file(filename):
    """Verifica si el archivo tiene una extensión permitida"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']

def save_upload_file(file, subfolder='uploads'):
    """Guarda un archivo subido y retorna la ruta"""
    if not file or file.filename == '':
        return None
    
    if not allowed_file(file.filename):
        raise ValueError('Tipo de archivo no permitido')
    
    # Crear carpeta si no existe
    upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], subfolder)
    os.makedirs(upload_folder, exist_ok=True)
    
    # Asegurar nombre de archivo
    filename = secure_filename(file.filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
    filename = timestamp + filename
    
    filepath = os.path.join(upload_folder, filename)
    file.save(filepath)
    
    return filepath

def delete_upload_file(filepath):
    """Elimina un archivo subido"""
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            return True
    except Exception as e:
        print(f"Error eliminando archivo: {e}")
    return False

# ============ ENVÍO DE CORREOS ============

def send_email_alert(recipient_email, subject, body, html_body=None):
    """Envía un email de alerta"""
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = current_app.config['MAIL_DEFAULT_SENDER']
        msg['To'] = recipient_email
        
        # Agregar versión texto
        part1 = MIMEText(body, 'plain')
        msg.attach(part1)
        
        # Agregar versión HTML si existe
        if html_body:
            part2 = MIMEText(html_body, 'html')
            msg.attach(part2)
        
        # Enviar
        with smtplib.SMTP(current_app.config['MAIL_SERVER'], current_app.config['MAIL_PORT']) as server:
            if current_app.config['MAIL_USE_TLS']:
                server.starttls()
            if current_app.config['MAIL_USERNAME']:
                server.login(current_app.config['MAIL_USERNAME'], current_app.config['MAIL_PASSWORD'])
            server.send_message(msg)
        
        return True
    except Exception as e:
        print(f"Error enviando email: {e}")
        return False

def send_certification_expiration_alert(certification, days_until_expiration):
    """Envía alerta de vencimiento de certificación"""
    user = certification.responsible
    
    subject = f"ALERTA: Certificación '{certification.name}' vence en {days_until_expiration} días"
    
    body = f"""
Estimado/a {user.full_name},

Le informamos que la certificación {certification.name} ({certification.norm})
vencerá en {days_until_expiration} días.

Detalles:
- Nombre: {certification.name}
- Norma: {certification.norm}
- Entidad Emisora: {certification.issuing_entity}
- Fecha de Vencimiento: {certification.expiration_date.strftime('%d/%m/%Y')}

Por favor, tome las acciones necesarias para la renovación.

Sistema de Gestión de Cumplimiento Normativo
Agroindustria Frutos de Oro S.A.C.
"""
    
    html_body = f"""
<html>
<body style="font-family: Arial, sans-serif;">
    <h2>ALERTA DE VENCIMIENTO DE CERTIFICACIÓN</h2>
    <p>Estimado/a <b>{user.full_name}</b>,</p>
    <p>Le informamos que la certificación <b>{certification.name}</b> vencerá en <b>{days_until_expiration} días</b>.</p>
    <h3>Detalles:</h3>
    <ul>
        <li><b>Nombre:</b> {certification.name}</li>
        <li><b>Norma:</b> {certification.norm}</li>
        <li><b>Entidad Emisora:</b> {certification.issuing_entity}</li>
        <li><b>Fecha de Vencimiento:</b> {certification.expiration_date.strftime('%d/%m/%Y')}</li>
    </ul>
    <p><b>Por favor, tome las acciones necesarias para la renovación.</b></p>
    <hr>
    <p><i>Sistema de Gestión de Cumplimiento Normativo - Agroindustria Frutos de Oro S.A.C.</i></p>
</body>
</html>
"""
    
    return send_email_alert(user.email, subject, body, html_body)

# ============ GENERACIÓN DE REPORTES ============

def generate_pdf_report(report_type, data):
    """Genera un reporte en PDF"""
    from app.models import Certification, Audit
    
    # Crear buffer en memoria
    buffer = io.BytesIO()
    
    # Crear documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=50,
        leftMargin=50,
        topMargin=50,
        bottomMargin=50
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilo personalizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a472a'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    # Título del reporte
    if report_type == 'certifications':
        elements.append(Paragraph('Reporte de Certificaciones', title_style))
        elements.append(Spacer(1, 0.3 * 72))
        
        # Datos de certificaciones
        cert_data = [['Certificación', 'Norma', 'Emisor', 'Vencimiento', 'Estado']]
        
        for cert in data:
            cert.get_status()
            cert_data.append([
                cert.name,
                cert.norm,
                cert.issuing_entity,
                cert.expiration_date.strftime('%d/%m/%Y'),
                cert.status.upper()
            ])
        
        # Crear tabla
        table = Table(cert_data, colWidths=[120, 80, 120, 100, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a472a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        
        elements.append(table)
    
    elif report_type == 'audits':
        elements.append(Paragraph('Reporte de Auditorías', title_style))
        elements.append(Spacer(1, 0.3 * 72))
        
        # Datos de auditorías
        audit_data = [['Tipo', 'Área', 'Fecha Programada', 'Estado', 'Hallazgos']]
        
        for audit in data:
            critical_findings = sum(1 for f in audit.findings if f.severity == 'critica')
            audit_data.append([
                audit.audit_type.upper(),
                audit.evaluated_area,
                audit.scheduled_date.strftime('%d/%m/%Y'),
                audit.status.upper(),
                str(critical_findings)
            ])
        
        table = Table(audit_data, colWidths=[80, 120, 120, 100, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a472a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        
        elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Guardar en archivo temporal
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], f'reporte_{report_type}_{timestamp}.pdf')
    
    with open(filepath, 'wb') as f:
        f.write(buffer.getvalue())
    
    return filepath

def generate_excel_report(report_type, data):
    """Genera un reporte en Excel"""
    wb = Workbook()
    ws = wb.active
    
    # Estilos
    header_fill = PatternFill(start_color='1a472a', end_color='1a472a', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    if report_type == 'certifications':
        ws.title = 'Certificaciones'
        ws.append(['Certificación', 'Norma', 'Emisor', 'Fecha Emisión', 'Fecha Vencimiento', 'Estado', 'Responsable'])
        
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        
        for cert in data:
            cert.get_status()
            ws.append([
                cert.name,
                cert.norm,
                cert.issuing_entity,
                cert.emission_date.strftime('%d/%m/%Y'),
                cert.expiration_date.strftime('%d/%m/%Y'),
                cert.status,
                cert.responsible.full_name
            ])
    
    elif report_type == 'audits':
        ws.title = 'Auditorías'
        ws.append(['Tipo', 'Área Evaluada', 'Responsable', 'Fecha Programada', 'Fecha Ejecución', 'Estado', 'Hallazgos Críticos'])
        
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        
        for audit in data:
            critical_count = sum(1 for f in audit.findings if f.severity == 'critica')
            ws.append([
                audit.audit_type,
                audit.evaluated_area,
                audit.responsible.full_name,
                audit.scheduled_date.strftime('%d/%m/%Y'),
                audit.executed_date.strftime('%d/%m/%Y') if audit.executed_date else '',
                audit.status,
                critical_count
            ])
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en archivo temporal
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], f'reporte_{report_type}_{timestamp}.xlsx')
    
    wb.save(filepath)
    return filepath

# ============ VERIFICACIÓN DE ALERTAS ============

def check_certification_expiration_alerts():
    """Verifica certificaciones próximas a vencer y envía alertas"""
    from app.models import Certification, Alert
    from datetime import timedelta
    
    today = datetime.now().date()
    
    # Verificar certificaciones próximas a vencer
    certifications = Certification.query.all()
    
    for cert in certifications:
        days_to_expiration = cert.days_to_expiration()
        
        # Alerta a 60 días
        if days_to_expiration == 60 and not cert.alert_sent_60:
            send_certification_expiration_alert(cert, 60)
            cert.alert_sent_60 = True
            create_alert(
                alert_type='expiration',
                related_id=cert.id,
                title=f'Certificación próxima a vencer: {cert.name}',
                message=f'Vencimiento en 60 días',
                severity='info',
                recipient_email=cert.responsible.email
            )
        
        # Alerta a 30 días
        elif days_to_expiration == 30 and not cert.alert_sent_30:
            send_certification_expiration_alert(cert, 30)
            cert.alert_sent_30 = True
            create_alert(
                alert_type='expiration',
                related_id=cert.id,
                title=f'Certificación próxima a vencer: {cert.name}',
                message=f'Vencimiento en 30 días',
                severity='warning',
                recipient_email=cert.responsible.email
            )
        
        # Alerta a 15 días
        elif days_to_expiration == 15 and not cert.alert_sent_15:
            send_certification_expiration_alert(cert, 15)
            cert.alert_sent_15 = True
            create_alert(
                alert_type='expiration',
                related_id=cert.id,
                title=f'URGENTE: Certificación próxima a vencer: {cert.name}',
                message=f'Vencimiento en 15 días - Acción requerida',
                severity='critical',
                recipient_email=cert.responsible.email
            )

def create_alert(alert_type, related_id, title, message, severity='info', recipient_email=None):
    """Crea una alerta en el sistema"""
    from app.models import Alert
    
    alert = Alert(
        alert_type=alert_type,
        related_id=related_id,
        title=title,
        message=message,
        severity=severity,
        recipient_email=recipient_email
    )
    
    from app import db
    db.session.add(alert)
    db.session.commit()
    
    return alert
